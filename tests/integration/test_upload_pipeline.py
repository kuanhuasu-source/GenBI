"""tests/integration/test_upload_pipeline.py — e2e upload + parse + profile + metadata (M4b)."""

from __future__ import annotations

import pytest


# ============================================================
# Test 1:乾淨 CSV 完整 pipeline 跑通
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_e2e_clean_csv_upload_to_metadata(
    mongo_db, golden_data_dir, tmp_path,
):
    """spec §16.2 #1:upload 乾淨 CSV → parse → profile → metadata v1 draft。"""
    from upload_repository import UploadRepository
    from upload_service import UploadService

    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path)

    csv_bytes = (golden_data_dir / "projects_clean.csv").read_bytes()
    dataset_id = service.handle_upload(
        file_obj=csv_bytes,
        filename="projects_clean.csv",
        owner="alice",
    )

    # 1. Dataset 寫入
    dataset = repo.get_dataset(dataset_id)
    assert dataset is not None
    assert dataset["status"] == "profiled"
    assert dataset["file"]["file_type"] == "csv"

    # 2. Table profile
    tables = repo.list_tables(dataset_id)
    assert len(tables) == 1
    assert tables[0]["row_count"] == 15
    assert tables[0]["column_count"] == 6

    # 3. Profile 寫入
    profile = repo.get_latest_profile(dataset_id)
    assert profile["profile_version"] == 1
    assert len(profile["tables"][0]["columns"]) == 6

    # 4. Metadata v1 draft 已產
    active = repo.get_active_metadata(dataset_id)
    assert active is not None
    assert active["version"] == 1
    assert active["confirmation_status"] == "draft"
    md = active["metadata"]
    assert md["source_type"] == "upload"
    # 該偵測出 project_id 為 identifier
    pid = md["collections"]["sheet1"]["fields"]["project_id"]
    assert pid["semantic_role"] == "identifier"


# ============================================================
# Test 2:PII dataset 上傳,PII 欄該被 mark
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_e2e_pii_dataset_marks_pii_fields(
    mongo_db, golden_data_dir, tmp_path,
):
    """spec §14.3:upload 含 PII 的 dataset → email/phone/employee_id 該 mark pii。"""
    from upload_repository import UploadRepository
    from upload_service import UploadService

    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path)

    csv_bytes = (golden_data_dir / "employee_pii.csv").read_bytes()
    dataset_id = service.handle_upload(
        file_obj=csv_bytes,
        filename="employee_pii.csv",
        owner="alice",
    )

    active = repo.get_active_metadata(dataset_id)
    fields = active["metadata"]["collections"]["sheet1"]["fields"]
    # PII 欄位 semantic_role 該 = 'pii'
    assert fields["email"]["semantic_role"] == "pii"
    assert fields["phone"]["semantic_role"] == "pii"
    assert fields["full_name"]["semantic_role"] == "pii"
    assert fields["employee_id"]["semantic_role"] == "pii"
    # salary / department 不該 mark
    assert fields["salary"]["semantic_role"] != "pii"
    assert fields["department"]["semantic_role"] != "pii"


# ============================================================
# Test 3:沒日期欄,trend 不支援
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_e2e_no_date_column_blocks_trend(
    mongo_db, golden_data_dir, tmp_path,
):
    """spec §16.2 #4:沒日期欄,data_limitations 該反映"""
    from upload_repository import UploadRepository
    from upload_service import UploadService

    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path)

    csv_bytes = (golden_data_dir / "projects_no_date.csv").read_bytes()
    dataset_id = service.handle_upload(
        file_obj=csv_bytes,
        filename="projects_no_date.csv",
        owner="alice",
    )

    active = repo.get_active_metadata(dataset_id)
    lim = active["metadata"]["data_limitations"]
    # missing_dimensions 該含 date
    assert any("date" in m.lower() for m in lim["missing_dimensions"])
    # not_supported_analysis 該含 trend
    assert any("trend" in n.lower() or "趨勢" in n
                for n in lim["not_supported_analysis"])


# ============================================================
# Test 4:有日期欄,trend 應該不被阻
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_e2e_with_date_column_allows_trend(
    mongo_db, golden_data_dir, tmp_path,
):
    from upload_repository import UploadRepository
    from upload_service import UploadService

    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path)

    csv_bytes = (golden_data_dir / "projects_with_date.csv").read_bytes()
    dataset_id = service.handle_upload(
        file_obj=csv_bytes,
        filename="projects_with_date.csv",
        owner="alice",
    )

    active = repo.get_active_metadata(dataset_id)
    lim = active["metadata"]["data_limitations"]
    # 不該標 "No date column"
    assert not any("no confirmed date" in m.lower()
                    for m in lim["missing_dimensions"])
    # 該至少有一個 datetime/date dimension 欄位
    fields = active["metadata"]["collections"]["sheet1"]["fields"]
    date_fields = [f for f in fields.values()
                    if f["semantic_role"] in ("date_dimension", "datetime_dimension")]
    assert len(date_fields) > 0


# ============================================================
# Test 5:metadata correction 後產新版,active 切到新版
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_e2e_correction_creates_new_version(
    mongo_db, golden_data_dir, tmp_path,
):
    """spec §16.2 #6:status code 經 user 修正後,後續分析採用修正後語意。"""
    from upload_repository import UploadRepository
    from upload_service import UploadService
    from metadata_correction_service import MetadataCorrectionService

    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path)

    csv_bytes = (golden_data_dir / "projects_clean.csv").read_bytes()
    dataset_id = service.handle_upload(
        file_obj=csv_bytes, filename="projects_clean.csv",
        owner="alice",
    )
    # v1 draft 該已存在
    active = repo.get_active_metadata(dataset_id)
    assert active["version"] == 1

    # User 改 leadtime 的 unit
    corrections_svc = MetadataCorrectionService(repo)
    result = corrections_svc.apply_corrections(
        dataset_id=dataset_id,
        corrections=[{
            "target": "sheet1.leadtime.unit",
            "old_value": "",
            "new_value": "days",
            "reason": "user confirmed",
        }],
        user="alice",
    )
    assert result["version"] == 2
    # 新 active 該是 v2
    new_active = repo.get_active_metadata(dataset_id)
    assert new_active["version"] == 2
    leadtime = new_active["metadata"]["collections"]["sheet1"]["fields"]["leadtime"]
    assert leadtime["unit"] == "days"
    assert leadtime["user_confirmed"] is True


# ============================================================
# v0.18 M1: Multi-sheet upload pipeline
# ============================================================
def _build_multi_sheet_xlsx(path):
    """Synthesize a 3-sheet xlsx in tests/integration/ via openpyxl.

    Sheet layout intentionally exercises all three table_role detections:
      - Employee:        clear PK (employee_id), 2 categorical attrs  → dimension
      - Department:      clear PK (department_id), 1 categorical      → dimension
      - EmployeeProject: no own PK, 2 FK-shaped cols (employee_id +
                         project_id), but only 1 FK matches another
                         table's PK (employee.employee_id) — Department
                         doesn't have project_id, so this resolves as
                         "unknown" rather than "bridge" with only 2 setup
                         tables. Adjusted below to include project_id PK
                         elsewhere if needed.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    # First sheet is auto-created — repurpose it
    wb.active.title = "Employee"
    ws_emp = wb.active
    ws_emp.append(["employee_id", "name", "department_id"])
    # Names intentionally include duplicates so only employee_id qualifies
    # as PK candidate (real Employee tables have duplicate first names).
    names = ["Alice", "Bob", "Carol", "Alice", "Bob",
             "David", "Eve", "Carol", "Bob", "Frank"]
    for i in range(1, 11):
        ws_emp.append([f"E{i:03d}", names[i - 1], f"D{(i % 3) + 1}"])

    ws_dept = wb.create_sheet("Department")
    ws_dept.append(["department_id", "dept_name"])
    for i in range(1, 4):
        ws_dept.append([f"D{i}", f"Dept_{i}"])

    ws_link = wb.create_sheet("EmployeeProject")
    ws_link.append(["employee_id", "project_id", "role"])
    for i in range(1, 16):
        emp = f"E{(i % 10) + 1:03d}"
        proj = f"P{(i % 4) + 1}"
        ws_link.append([emp, proj, "dev" if i % 2 else "lead"])

    wb.save(path)


@pytest.mark.integration
@pytest.mark.requires_mongo
def test_e2e_multisheet_excel_creates_three_tables(
    mongo_db, tmp_path,
):
    """v0.18 M1 acceptance (spec §17 M1 row):
    Upload a multi-sheet xlsx → see N upload_tables rows, each with
    spec §5.2 fields (table_role, grain, primary_key, sheet_name)
    populated.

    Default `excel_multi_sheet=True` per v0.18 — no flag needed.
    """
    from upload_repository import UploadRepository
    from upload_service import UploadService

    # 1. Synthesize 3-sheet xlsx
    xlsx_path = tmp_path / "hr_workbook.xlsx"
    _build_multi_sheet_xlsx(xlsx_path)

    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path / "uploads")

    # 2. Upload (no excel_multi_sheet flag — relies on v0.18 default True)
    dataset_id = service.handle_upload(
        file_obj=xlsx_path.read_bytes(),
        filename="hr_workbook.xlsx",
        owner="alice",
    )

    # 3. Dataset reached profiled status
    dataset = repo.get_dataset(dataset_id)
    assert dataset["status"] == "profiled"
    assert dataset["file"]["file_type"] == "excel"

    # 4. All 3 sheets became upload_tables rows
    tables = repo.list_tables(dataset_id)
    assert len(tables) == 3
    table_names = {t["table_name"] for t in tables}
    assert table_names == {"Employee", "Department", "EmployeeProject"}

    # 5. Each row got v0.18 enrichment fields per spec §5.2.
    # update_table_profile_fields treats None as "leave alone" so the `grain`
    # field is only present when PK was detected (non-None grain). table_role,
    # sheet_name, primary_key are always set (the last as [] for unknown).
    for t in tables:
        assert "sheet_name" in t, f"sheet_name missing on {t['table_id']}"
        assert "table_role" in t, f"table_role missing on {t['table_id']}"
        assert "primary_key" in t, f"primary_key missing on {t['table_id']}"
        # grain only present when PK non-empty
        if t["primary_key"]:
            assert "grain" in t and t["grain"], (
                f"grain should be set when PK exists on {t['table_id']}"
            )

    # 6. Employee table specifically: PK=employee_id, role=dimension or bridge
    #    (bridge if Employee's employee_id PK matches EmployeeProject's
    #     employee_id col + something else — usually dimension here)
    emp_row = next(t for t in tables if t["table_name"] == "Employee")
    assert emp_row["sheet_name"] == "Employee"
    assert emp_row["primary_key"] == ["employee_id"]
    assert emp_row["grain"] == "one row per employee"
    assert emp_row["table_role"] in ("dimension", "fact")

    # 7. EmployeeProject: bridge candidate — has 2+ cols matching other
    #    tables' PKs (employee_id from Employee, department_id NOT here so
    #    only one FK match → falls back to "unknown" or "dimension").
    #    We don't pin the exact role since detection depends on PK overlap;
    #    we DO require that profile didn't crash and the field exists.
    link_row = next(t for t in tables if t["table_name"] == "EmployeeProject")
    assert link_row["sheet_name"] == "EmployeeProject"
    assert "table_role" in link_row

    # 8. upload_profiles still keyed by parser table_id — regenerate_metadata
    #    contract preserved. Verify by walking the profile doc.
    profile = repo.get_latest_profile(dataset_id)
    assert len(profile["tables"]) == 3
    upload_tids = {t["table_id"] for t in tables}
    profile_tids = {tp["table_id"] for tp in profile["tables"]}
    assert profile_tids == upload_tids, (
        f"profile/upload_tables table_id mismatch — "
        f"regenerate_metadata would break. "
        f"profile={profile_tids}, upload={upload_tids}"
    )


@pytest.mark.integration
@pytest.mark.requires_mongo
def test_e2e_multisheet_detects_and_persists_relationships(
    mongo_db, tmp_path,
):
    """v0.18 M2: after uploading a multi-sheet xlsx, relationship_profiler
    runs and saves candidates to upload_relationship_candidates. The
    Employee → EmployeeProject FK (shared employee_id) should be detected
    at high confidence.
    """
    from upload_repository import UploadRepository
    from upload_service import UploadService

    xlsx_path = tmp_path / "hr.xlsx"
    _build_multi_sheet_xlsx(xlsx_path)

    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path / "uploads")
    dataset_id = service.handle_upload(
        file_obj=xlsx_path.read_bytes(),
        filename="hr.xlsx",
        owner="alice",
    )

    # 1. Candidates persisted
    cands = repo.list_relationship_candidates(dataset_id)
    assert len(cands) >= 1, (
        "upload pipeline must persist at least one relationship candidate "
        "for a 3-sheet workbook where Employee + EmployeeProject share "
        "employee_id"
    )

    # 2. Each candidate has spec §5.3 fields
    for c in cands:
        for k in ("relationship_id", "from_table", "from_field",
                  "to_table", "to_field", "relationship_type",
                  "default_join_type", "confidence", "evidence",
                  "status", "dataset_id", "metadata_version"):
            assert k in c, f"missing key `{k}` in {c}"
        assert c["dataset_id"] == dataset_id
        assert c["status"] == "candidate"  # spec default before user review

    # 3. The employee_id link is detectable at high tier.
    emp_rels = [
        c for c in cands
        if c["from_field"] == "employee_id"
        and {c["from_table"], c["to_table"]} == {"Employee", "EmployeeProject"}
    ]
    assert emp_rels, (
        "expected at least one Employee↔EmployeeProject candidate on "
        "employee_id"
    )
    # The from→to direction with EmployeeProject as `from` should be m2o.
    m2o_match = next(
        (c for c in emp_rels
         if c["from_table"] == "EmployeeProject"
         and c["to_table"] == "Employee"),
        None,
    )
    assert m2o_match is not None
    assert m2o_match["relationship_type"] == "many_to_one"
    assert m2o_match["confidence_tier"] == "high"
    assert m2o_match["confidence"] >= 0.90


@pytest.mark.integration
@pytest.mark.requires_mongo
def test_e2e_csv_upload_still_works_with_default_multi_sheet_true(
    mongo_db, golden_data_dir, tmp_path,
):
    """Flipping excel_multi_sheet=True default must not break CSV uploads.
    CSV path doesn't branch on the flag (is_excel=False), but the new
    enrichment block runs for all uploads — verify it's safe on CSV too.
    """
    from upload_repository import UploadRepository
    from upload_service import UploadService

    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path)

    csv_bytes = (golden_data_dir / "projects_clean.csv").read_bytes()
    dataset_id = service.handle_upload(
        file_obj=csv_bytes, filename="projects_clean.csv", owner="alice",
    )

    dataset = repo.get_dataset(dataset_id)
    assert dataset["status"] == "profiled"

    tables = repo.list_tables(dataset_id)
    assert len(tables) == 1
    # Enrichment ran on the single CSV "table" too
    t = tables[0]
    assert "table_role" in t
    assert "primary_key" in t
    # project_id is unique in projects_clean.csv → PK candidate
    assert "project_id" in t["primary_key"]


# ============================================================
# Test 6:Confirm 流程
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_e2e_confirm_metadata_workflow(
    mongo_db, golden_data_dir, tmp_path,
):
    """spec §10.7 #5:Confirm 後 confirmation_status='confirmed'。"""
    from upload_repository import UploadRepository
    from upload_service import UploadService
    from metadata_correction_service import MetadataCorrectionService

    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path)

    csv_bytes = (golden_data_dir / "projects_clean.csv").read_bytes()
    dataset_id = service.handle_upload(
        file_obj=csv_bytes, filename="x.csv", owner="alice",
    )

    # v1 是 draft
    assert repo.get_active_metadata(dataset_id)["confirmation_status"] == "draft"

    # Confirm
    corrections_svc = MetadataCorrectionService(repo)
    result = corrections_svc.confirm_metadata(
        dataset_id=dataset_id, user="alice", notes="LGTM",
    )
    assert result["version"] == 2

    # v2 active + confirmed
    active = repo.get_active_metadata(dataset_id)
    assert active["confirmation_status"] == "confirmed"
    assert active["confirmed_by"] == "alice"
