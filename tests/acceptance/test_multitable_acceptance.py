"""tests/acceptance/test_multitable_acceptance.py — v0.18 M1/M7

Spec §14.5 lists 12 acceptance tests for the multi-table upgrade. This file
is the M7 scaffold — implements 3 of those 12 (the ones whose dependencies
are landed in M1). Future PRs (M2-M6) add the rest:

| #  | name                                                 | status     | needs |
|----|------------------------------------------------------|------------|-------|
| 1  | test_multisheet_1_parse_all_sheets                   | ✅ here    | M1    |
| 2  | test_multisheet_2_profile_each_sheet                 | ✅ here    | M1    |
| 3  | test_multisheet_3_infer_high_confidence_relationship | ✅ here    | M2    |
| 4  | test_multisheet_4_low_confidence_requires_review     | ✅ here    | M2    |
| 5  | test_multisheet_5_confirm_relationship_creates_..    | ✅ here    | M7    |
| 6  | test_multisheet_6_unconfirmed_relationship_not_used  | ✅ here    | M4    |
| 7  | test_multisheet_7_duckdb_join_confirmed_relationship | ✅ here    | M4    |
| 8  | test_multisheet_8_add_calculated_column              | ✅ here    | M5    |
| 9  | test_multisheet_9_create_derived_table               | ✅ here    | M5    |
| 10 | test_multisheet_10_visualize_derived_table           | ✅ here    | M5    |
| 11 | test_multisheet_11_save_analysis_template_lineage    | ✅ here    | M6    |
| 12 | test_multisheet_12_schema_driven_baseline_unchanged  | ✅ here    | always|

Test 12 is the frozen-baseline rule from spec §14.1: schema-driven path must
not change. The proxy check here is "all schema-driven critical files compile
cleanly and their public entry points still exist." A heavier check that
actually runs test_runner.py --domain tflex is in scripts/run_regression_gate.py
--mode full (skipped if Ollama/MongoDB unavailable).
"""

from __future__ import annotations

from pathlib import Path

import pytest


# ============================================================
# Fixture helper — synthesize a 3-sheet xlsx in tmp_path.
#
# Sheet layout intentionally exercises the M1 profile pipeline:
#   - Employee:        clear PK (employee_id), dim attrs
#   - Department:      clear PK (department_id), 1 attr
#   - EmployeeProject: FK columns matching both, no own PK
#
# Names include duplicates so only *_id columns qualify as PK
# candidates (realistic data shape — matches the rule in M1
# multi_table_profiler tests).
# ============================================================
def _build_3sheet_xlsx(path: Path) -> None:
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = "Employee"
    ws_emp = wb.active
    ws_emp.append(["employee_id", "name", "department_id"])
    names = ["Alice", "Bob", "Carol", "Alice", "Bob",
             "David", "Eve", "Carol", "Bob", "Frank"]
    for i in range(1, 11):
        ws_emp.append([f"E{i:03d}", names[i - 1], f"D{(i % 3) + 1}"])

    ws_dept = wb.create_sheet("Department")
    ws_dept.append(["department_id", "dept_name"])
    for i in range(1, 4):
        ws_dept.append([f"D{i}", f"Dept_{i}"])

    ws_link = wb.create_sheet("EmployeeProject")
    ws_link.append(["employee_id", "project_id", "role"])
    for i in range(1, 16):
        emp = f"E{(i % 10) + 1:03d}"
        proj = f"P{(i % 4) + 1}"
        ws_link.append([emp, proj, "dev" if i % 2 else "lead"])

    wb.save(path)


# ============================================================
# Test 1 (spec §14.5 #1): parse_all_sheets
#   Upload an N-sheet Excel workbook → upload_tables has N rows,
#   each backed by a parquet file. v0.18 default excel_multi_sheet=True.
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_multisheet_1_parse_all_sheets(mongo_db, tmp_path):
    from upload_repository import UploadRepository
    from upload_service import UploadService

    xlsx_path = tmp_path / "wb.xlsx"
    _build_3sheet_xlsx(xlsx_path)

    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path / "uploads")
    dataset_id = service.handle_upload(
        file_obj=xlsx_path.read_bytes(),
        filename="wb.xlsx",
        owner="alice",
    )

    # All 3 sheets parsed into upload_tables (not just the first one).
    tables = repo.list_tables(dataset_id)
    assert len(tables) == 3, (
        f"expected 3 tables from 3-sheet xlsx, got {len(tables)}. "
        f"excel_multi_sheet=True must be the default."
    )

    # Each row points at a real parquet file (storage.format / .path).
    for t in tables:
        assert t["storage"]["format"] == "parquet"
        parquet_path = Path(t["storage"]["path"])
        assert parquet_path.exists(), (
            f"parquet file missing for table {t['table_id']}: {parquet_path}"
        )
        # Each table has non-zero row + col counts (i.e., parser didn't
        # silently produce an empty stub).
        assert t["row_count"] > 0
        assert t["column_count"] > 0


# ============================================================
# Test 2 (spec §14.5 #2): profile_each_sheet
#   After upload, profile_multi_table enrichment fields must be on each
#   upload_tables row + column-level profile must be in upload_profiles.
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_multisheet_2_profile_each_sheet(mongo_db, tmp_path):
    from upload_repository import UploadRepository
    from upload_service import UploadService

    xlsx_path = tmp_path / "wb.xlsx"
    _build_3sheet_xlsx(xlsx_path)

    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path / "uploads")
    dataset_id = service.handle_upload(
        file_obj=xlsx_path.read_bytes(),
        filename="wb.xlsx",
        owner="alice",
    )

    # 1. Column-level profile: upload_profiles doc has 3 entries.
    profile = repo.get_latest_profile(dataset_id)
    assert profile is not None
    assert len(profile["tables"]) == 3
    for tp in profile["tables"]:
        assert "columns" in tp
        assert len(tp["columns"]) > 0
        # spec §7.1 minimums on every column
        for col in tp["columns"]:
            assert "name" in col
            assert "physical_type" in col
            assert "null_count" in col

    # 2. Table-level enrichment per spec §5.2 fields populated on each
    # upload_tables row (the M1 deliverable).
    tables = repo.list_tables(dataset_id)
    for t in tables:
        assert "sheet_name" in t, (
            f"sheet_name missing on table {t['table_id']} — multi-table "
            f"enrichment didn't run"
        )
        assert "table_role" in t
        assert "primary_key" in t  # may be [] for unknown role

    # 3. Employee specifically: PK detected, grain phrased correctly.
    emp = next(t for t in tables if t["table_name"] == "Employee")
    assert emp["primary_key"] == ["employee_id"]
    assert emp["grain"] == "one row per employee"
    assert emp["table_role"] in ("dimension", "fact")


# ============================================================
# Test 3 (spec §14.5 #3): infer_high_confidence_relationship
#   3-sheet xlsx with shared employee_id PK/FK → ≥0.90 confidence
#   m2o relationship detected + persisted at status='candidate'.
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_multisheet_3_infer_high_confidence_relationship(mongo_db, tmp_path):
    from upload_repository import UploadRepository
    from upload_service import UploadService

    xlsx_path = tmp_path / "wb.xlsx"
    _build_3sheet_xlsx(xlsx_path)
    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path / "u")
    dataset_id = service.handle_upload(
        file_obj=xlsx_path.read_bytes(),
        filename="wb.xlsx", owner="alice",
    )

    cands = repo.list_relationship_candidates(dataset_id)
    # EmployeeProject.employee_id (low unique) → Employee.employee_id (PK)
    # is the canonical m2o we expect at high tier.
    high_m2o = [
        c for c in cands
        if c["confidence_tier"] == "high"
        and c["relationship_type"] == "many_to_one"
    ]
    assert high_m2o, (
        "expected at least one high-confidence many_to_one relationship "
        "from a 3-sheet xlsx with shared keys"
    )
    # Default status is candidate awaiting review (spec §5.3).
    assert all(c["status"] == "candidate" for c in high_m2o)


# ============================================================
# Test 4 (spec §14.5 #4): low_confidence_requires_review
#   2-sheet xlsx with partial-overlap shared key → confidence falls in
#   review_required tier (0.70-0.89) and status stays 'candidate' so the
#   user must explicitly confirm before it can be used for join.
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_multisheet_4_low_confidence_requires_review(mongo_db, tmp_path):
    import openpyxl
    from upload_repository import UploadRepository
    from upload_service import UploadService

    # Synthesize a workbook where shared key has only partial overlap.
    xlsx_path = tmp_path / "weak.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "orders"
    ws = wb.active
    ws.append(["order_id", "customer_id"])
    # 10 orders, customer_id covers C1..C5 (5 distinct, not unique)
    for i in range(1, 11):
        ws.append([f"O{i:03d}", f"C{(i % 5) + 1}"])

    ws2 = wb.create_sheet("customers")
    ws2.append(["customer_id", "name"])
    # PK side: only C1..C3 present — orders refers to C4/C5 too →
    # overlap_ratio < 1.0 → confidence in review tier.
    for cid, nm in [("C1", "Alice"), ("C2", "Bob"), ("C3", "Carol")]:
        ws2.append([cid, nm])

    wb.save(xlsx_path)

    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path / "u")
    dataset_id = service.handle_upload(
        file_obj=xlsx_path.read_bytes(),
        filename="weak.xlsx", owner="alice",
    )

    cands = repo.list_relationship_candidates(dataset_id)
    # The orders → customers candidate should exist but NOT be high tier.
    o2c = next(
        (c for c in cands
         if c["from_table"] == "orders" and c["to_table"] == "customers"
         and c["from_field"] == "customer_id"),
        None,
    )
    assert o2c is not None, "expected partial-overlap candidate to be saved"
    # Spec §14.5 #4 intent: partial overlap → tier 'review_required' or
    # 'weak'; never auto-high.
    assert o2c["confidence_tier"] in ("review_required", "weak"), (
        f"partial-overlap rel should require review, got tier="
        f"{o2c['confidence_tier']} confidence={o2c['confidence']}"
    )
    # And status stays 'candidate' — must NOT auto-confirm.
    assert o2c["status"] == "candidate"


# ============================================================
# Test 5 (spec §14.5 #5): confirm_relationship_creates_metadata_version
#   Upload 3-sheet → confirm a relationship → confirm_metadata →
#   new metadata version exists with the confirmed relationship
#   projected into metadata.relationships. Spec §9.2 + §18 rule 27.
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_multisheet_5_confirm_relationship_creates_metadata_version(
    mongo_db, tmp_path,
):
    from metadata_correction_service import MetadataCorrectionService
    from upload_repository import UploadRepository
    from upload_service import UploadService

    xlsx_path = tmp_path / "wb.xlsx"
    _build_3sheet_xlsx(xlsx_path)
    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path / "u")
    dataset_id = service.handle_upload(
        file_obj=xlsx_path.read_bytes(),
        filename="wb.xlsx", owner="alice",
    )

    # Initial metadata version is a draft (auto-generated at upload).
    initial = repo.get_active_metadata(dataset_id)
    assert initial["confirmation_status"] == "draft"
    initial_version = initial["version"]

    # Find + confirm the high-confidence Employee↔EmployeeProject rel.
    cands = repo.list_relationship_candidates(dataset_id)
    target = next(
        c for c in cands
        if c["relationship_type"] == "many_to_one"
        and c["from_field"] == "employee_id"
        and {c["from_table"], c["to_table"]} == {
            "Employee", "EmployeeProject",
        }
    )
    repo.update_relationship_status(
        dataset_id, target["relationship_id"],
        status="confirmed", user="alice",
    )

    # Call confirm_metadata → should write a new version that contains
    # the confirmed relationship.
    correction_svc = MetadataCorrectionService(repo)
    result = correction_svc.confirm_metadata(
        dataset_id=dataset_id, user="alice",
        notes="LGTM after rel review",
    )
    assert result["version"] > initial_version
    assert result["already_confirmed"] is False
    assert result["n_relationships_merged"] >= 1

    # New active version has metadata.relationships with the confirmed rel.
    new_active = repo.get_active_metadata(dataset_id)
    assert new_active["confirmation_status"] == "confirmed"
    assert new_active["confirmed_by"] == "alice"
    assert "relationships" in new_active["metadata"]
    rels = new_active["metadata"]["relationships"]
    merged_ids = {r["relationship_id"] for r in rels}
    assert target["relationship_id"] in merged_ids

    # Spec §18 rule 27 — only executable fields, no evidence/confidence leak.
    confirmed_in_md = next(
        r for r in rels if r["relationship_id"] == target["relationship_id"]
    )
    assert confirmed_in_md["from_table"] == target["from_table"]
    assert confirmed_in_md["to_table"] == target["to_table"]
    assert confirmed_in_md["relationship_type"] == "many_to_one"
    assert "evidence" not in confirmed_in_md
    assert "confidence" not in confirmed_in_md


# ============================================================
# Test 6 (spec §14.5 #6): unconfirmed_relationship_not_used_in_join
#   Upload 3-sheet xlsx → relationships detected as candidates (not
#   confirmed). Attempting a JOIN must fail BEFORE execution with a
#   spec §14.6 anti-pattern error.
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_multisheet_6_unconfirmed_relationship_not_used_in_join(
    mongo_db, tmp_path,
):
    pytest.importorskip("duckdb")
    from upload_repository import UploadRepository
    from upload_service import UploadService
    from duckdb_engine import build_engine_for_dataset

    xlsx_path = tmp_path / "wb.xlsx"
    _build_3sheet_xlsx(xlsx_path)
    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path / "u")
    dataset_id = service.handle_upload(
        file_obj=xlsx_path.read_bytes(),
        filename="wb.xlsx", owner="alice",
    )

    # Candidates exist but none are confirmed.
    cands = repo.list_relationship_candidates(dataset_id)
    assert all(c["status"] == "candidate" for c in cands)

    engine = build_engine_for_dataset(repo, dataset_id)
    # The parser registers tables by table_id (lowercase). Sheets:
    # "Employee" → table_id "employee", same for the others.
    sql = (
        "SELECT employee.name, employeeproject.role "
        "FROM employee "
        "JOIN employeeproject "
        "ON employee.employee_id = employeeproject.employee_id"
    )
    result = engine.execute_safe_with_join_validation(
        sql, confirmed_relationships=cands,
    )
    engine.close()

    assert result.success is False, (
        "Spec §14.6 anti-pattern: unconfirmed relationship must NOT "
        "execute as a join"
    )
    assert result.error_type == "JoinNotConfirmed"
    assert "candidate" in result.error.lower()


# ============================================================
# Test 7 (spec §14.5 #7): duckdb_join_confirmed_relationship
#   Upload 3-sheet xlsx → user confirms Employee↔EmployeeProject
#   relationship → DuckDB JOIN executes + returns correct rows.
#   Verifies the spec §17 M4 acceptance criterion:
#     "可用 confirmed relationship 做跨 sheet join 並回傳正確聚合結果"
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_multisheet_7_duckdb_join_confirmed_relationship(
    mongo_db, tmp_path,
):
    pytest.importorskip("duckdb")
    from upload_repository import UploadRepository
    from upload_service import UploadService
    from duckdb_engine import build_engine_for_dataset

    xlsx_path = tmp_path / "wb.xlsx"
    _build_3sheet_xlsx(xlsx_path)
    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    service = UploadService(upload_repo=repo, uploads_root=tmp_path / "u")
    dataset_id = service.handle_upload(
        file_obj=xlsx_path.read_bytes(),
        filename="wb.xlsx", owner="alice",
    )

    # Find the Employee ↔ EmployeeProject candidate (m2o, high tier)
    cands = repo.list_relationship_candidates(dataset_id)
    target = next(
        c for c in cands
        if c["relationship_type"] == "many_to_one"
        and c["from_field"] == "employee_id"
        and {c["from_table"], c["to_table"]} == {
            "Employee", "EmployeeProject",
        }
    )

    # User confirms it.
    repo.update_relationship_status(
        dataset_id, target["relationship_id"],
        status="confirmed",
        user="alice",
    )

    # Re-list — now status should be confirmed for that one.
    cands_after = repo.list_relationship_candidates(dataset_id)
    confirmed = [c for c in cands_after if c["status"] == "confirmed"]
    assert len(confirmed) == 1

    # JOIN executes successfully and returns the expected row count.
    # EmployeeProject has 15 rows; each row's employee_id resolves to
    # exactly one Employee → 15-row join result.
    engine = build_engine_for_dataset(repo, dataset_id)
    sql = (
        "SELECT employee.name, employeeproject.role "
        "FROM employee "
        "JOIN employeeproject "
        "ON employee.employee_id = employeeproject.employee_id"
    )
    result = engine.execute_safe_with_join_validation(
        sql, confirmed_relationships=cands_after,
    )
    engine.close()

    assert result.success, (
        f"confirmed-relationship JOIN should execute; error: {result.error}"
    )
    assert len(result.df) == 15
    assert set(result.df.columns) == {"name", "role"}


# ============================================================
# Test 8 (spec §14.5 #8): add_calculated_column
#   Upload xlsx → start analysis session → add_column step adds a
#   calculated column → step persisted with output_schema reflecting
#   the new column + parquet written + queryable.
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_multisheet_8_add_calculated_column(mongo_db, tmp_path):
    from analysis_step_service import AnalysisStepService
    from upload_repository import UploadRepository
    from upload_service import UploadService

    xlsx_path = tmp_path / "wb.xlsx"
    _build_3sheet_xlsx(xlsx_path)
    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    uploads_root = tmp_path / "u"
    service = UploadService(upload_repo=repo, uploads_root=uploads_root)
    dataset_id = service.handle_upload(
        file_obj=xlsx_path.read_bytes(),
        filename="wb.xlsx", owner="alice",
    )

    # Mark metadata as confirmed so create_session has a version anchor.
    # (Without this, dataset.active_metadata_version may be None → step
    # service falls back to version 1, which is fine for this test.)
    svc = AnalysisStepService(upload_repo=repo, uploads_root=uploads_root)
    session_id = svc.create_session(dataset_id, owner="alice")

    # Step 1: add_column on EmployeeProject — derive a numeric flag col.
    s1 = svc.add_step(
        session_id, action_type="add_column",
        params={
            "input_table": "employeeproject",
            "new_column": "is_lead",
            # role is "lead" or "dev"; eval-style boolean → 0/1.
            # df.eval doesn't support `==` with string + python engine
            # cleanly; use len() trick.
            "formula": "role.str.len() == 4",
        },
        user_query="flag lead rows",
    )
    assert s1["status"] == "completed", (
        f"add_column step should succeed; error: {s1.get('error_message')}"
    )
    # output_schema records the new column
    col_names = {c["name"] for c in s1["output_schema"]}
    assert "is_lead" in col_names
    # parquet was written
    assert s1["storage"]["format"] == "parquet"
    # Resolve the derived table — `is_lead` column visible
    df = svc.resolve_table(session_id, s1["output_table"])
    assert "is_lead" in df.columns
    assert df["is_lead"].dtype == bool

    # Step doc persisted in MongoDB with full lineage
    step_doc = repo.get_analysis_step(s1["step_id"])
    assert step_doc is not None
    assert step_doc["input_tables"] == ["employeeproject"]
    assert step_doc["action_type"] == "add_column"


# ============================================================
# Test 9 (spec §14.5 #9): create_derived_table
#   Chain extract → add_column → create_table → derived table is
#   queryable by its user-supplied name + lineage preserved through
#   all 3 steps.
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_multisheet_9_create_derived_table(mongo_db, tmp_path):
    from analysis_step_service import AnalysisStepService
    from upload_repository import UploadRepository
    from upload_service import UploadService

    xlsx_path = tmp_path / "wb.xlsx"
    _build_3sheet_xlsx(xlsx_path)
    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    uploads_root = tmp_path / "u"
    service = UploadService(upload_repo=repo, uploads_root=uploads_root)
    dataset_id = service.handle_upload(
        file_obj=xlsx_path.read_bytes(),
        filename="wb.xlsx", owner="alice",
    )
    svc = AnalysisStepService(upload_repo=repo, uploads_root=uploads_root)
    session_id = svc.create_session(dataset_id, owner="alice")

    # 3-step chain.
    s1 = svc.add_step(
        session_id, action_type="extract_data",
        params={"input_table": "employee"},
        user_query="all employees",
    )
    s2 = svc.add_step(
        session_id, action_type="aggregate",
        params={
            "input_table": s1["output_table"],
            "group_by": ["department_id"],
            "aggregations": [
                {"column": "employee_id", "function": "count_distinct",
                 "alias": "headcount"},
            ],
        },
        user_query="headcount per dept",
    )
    s3 = svc.add_step(
        session_id, action_type="create_table",
        params={
            "input_table": s2["output_table"],
            "new_name": "dept_headcount",
        },
        user_query="save aggregation as named table",
    )

    # All 3 steps must succeed
    for s, name in [(s1, "extract"), (s2, "aggregate"), (s3, "create")]:
        assert s["status"] == "completed", (
            f"{name} step failed: {s.get('error_message')}"
        )

    # Derived table queryable by user-supplied name
    df = svc.resolve_table(session_id, "dept_headcount")
    assert "department_id" in df.columns
    assert "headcount" in df.columns
    assert len(df) >= 1

    # Lineage preserved through all 3 steps in correct order
    steps = repo.list_analysis_steps(session_id)
    assert len(steps) == 3
    assert [s["step_no"] for s in steps] == [1, 2, 3]
    assert [s["action_type"] for s in steps] == [
        "extract_data", "aggregate", "create_table",
    ]
    # Each step's input traces to the prior step's output
    assert steps[1]["input_tables"] == [steps[0]["output_table"]]
    assert steps[2]["input_tables"] == [steps[1]["output_table"]]
    # The final create_table step uses the user-supplied name
    assert steps[2]["output_table"] == "dept_headcount"


# ============================================================
# Test 10 (spec §14.5 #10): visualize_derived_table
#   A derived table from a prior step can be referenced as input to
#   a visualize step. Chart spec is persisted on the step doc.
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_multisheet_10_visualize_derived_table(mongo_db, tmp_path):
    from analysis_step_service import AnalysisStepService
    from upload_repository import UploadRepository
    from upload_service import UploadService

    xlsx_path = tmp_path / "wb.xlsx"
    _build_3sheet_xlsx(xlsx_path)
    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    uploads_root = tmp_path / "u"
    service = UploadService(upload_repo=repo, uploads_root=uploads_root)
    dataset_id = service.handle_upload(
        file_obj=xlsx_path.read_bytes(),
        filename="wb.xlsx", owner="alice",
    )
    svc = AnalysisStepService(upload_repo=repo, uploads_root=uploads_root)
    session_id = svc.create_session(dataset_id, owner="alice")

    # Build a derived table to visualize
    s1 = svc.add_step(
        session_id, action_type="aggregate",
        params={
            "input_table": "employee",
            "group_by": ["department_id"],
            "aggregations": [
                {"column": "employee_id", "function": "count",
                 "alias": "headcount"},
            ],
        },
    )
    assert s1["status"] == "completed"

    # Visualize step references the derived table
    s2 = svc.add_step(
        session_id, action_type="visualize",
        params={
            "input_table": s1["output_table"],
            "chart_type": "bar",
            "x": "department_id",
            "y": "headcount",
        },
        user_query="show headcount by dept as bar",
    )
    assert s2["status"] == "completed"

    # Chart spec persisted on the step doc (the visualize action doesn't
    # render — it records intent so M6 saved-chart can pick up later).
    spec = s2["chart_spec"]
    assert spec["chart_type"] == "bar"
    assert spec["x"] == "department_id"
    assert spec["y"] == "headcount"
    # visualize action doesn't materialize a parquet (no output data)
    assert "storage" not in s2

    # Lineage: visualize step input traces back to the aggregate step's
    # derived output.
    assert s2["input_tables"] == [s1["output_table"]]


# ============================================================
# Test 11 (spec §14.5 #11): save_analysis_template_lineage
#   Build a 3-step M5 chain → save as analysis_template via M6 →
#   verify spec §18 rule 27: asset binds dataset_id, metadata_version,
#   source_step_ids. Then verify the template payload preserves
#   replayable action+params + drift check works.
# ============================================================
@pytest.mark.integration
@pytest.mark.requires_mongo
def test_multisheet_11_save_analysis_template_lineage(mongo_db, tmp_path):
    from analysis_asset_service import AnalysisAssetService
    from analysis_step_service import AnalysisStepService
    from metadata_correction_service import MetadataCorrectionService
    from upload_repository import UploadRepository
    from upload_service import UploadService

    # 1. Upload + create session
    xlsx_path = tmp_path / "wb.xlsx"
    _build_3sheet_xlsx(xlsx_path)
    repo = UploadRepository(mongo_db)
    repo.ensure_indexes()
    uploads_root = tmp_path / "u"
    service = UploadService(upload_repo=repo, uploads_root=uploads_root)
    dataset_id = service.handle_upload(
        file_obj=xlsx_path.read_bytes(),
        filename="wb.xlsx", owner="alice",
    )
    step_svc = AnalysisStepService(upload_repo=repo, uploads_root=uploads_root)
    session_id = step_svc.create_session(dataset_id, owner="alice")

    # 2. Build 3-step chain that an analyst would actually want to template:
    #    extract → aggregate → visualize
    s1 = step_svc.add_step(
        session_id, action_type="extract_data",
        params={"input_table": "employee"},
        user_query="all employees",
    )
    s2 = step_svc.add_step(
        session_id, action_type="aggregate",
        params={
            "input_table": s1["output_table"],
            "group_by": ["department_id"],
            "aggregations": [
                {"column": "employee_id", "function": "count_distinct",
                 "alias": "headcount"},
            ],
        },
        user_query="headcount per dept",
    )
    s3 = step_svc.add_step(
        session_id, action_type="visualize",
        params={
            "input_table": s2["output_table"],
            "chart_type": "bar",
            "x": "department_id",
            "y": "headcount",
        },
        user_query="bar chart of headcount",
    )
    for s, n in [(s1, "extract"), (s2, "agg"), (s3, "viz")]:
        assert s["status"] == "completed", f"{n} step failed"

    # 3. Save the chain as an analysis_template via M6
    correction_svc = MetadataCorrectionService(repo)
    asset_svc = AnalysisAssetService(repo, correction_svc)
    step_ids = [s1["step_id"], s2["step_id"], s3["step_id"]]
    asset_id = asset_svc.save_template_from_steps(
        session_id=session_id,
        step_ids=step_ids,
        name="dept_headcount_template",
        description="reusable headcount-per-dept analysis",
        user="alice",
    )

    # 4. Spec §18 rule 27: asset must bind dataset_id, metadata_version,
    # source_step_ids.
    asset = repo.get_asset(asset_id)
    assert asset is not None
    assert asset["asset_type"] == "analysis_template"
    assert asset["dataset_id"] == dataset_id
    assert "metadata_version" in asset and asset["metadata_version"] >= 1
    assert asset["source_step_ids"] == step_ids

    # 5. The payload is replayable — each step's action_type + params
    # preserved in step_no order.
    payload = asset["asset_payload"]
    assert payload["n_steps"] == 3
    steps = payload["steps"]
    assert [s["step_no"] for s in steps] == [1, 2, 3]
    assert [s["action_type"] for s in steps] == [
        "extract_data", "aggregate", "visualize",
    ]
    # Aggregation params preserved (needed for replay).
    agg_step = steps[1]
    assert agg_step["params"]["group_by"] == ["department_id"]
    assert (
        agg_step["params"]["aggregations"][0]["function"]
        == "count_distinct"
    )
    # Visualize chart spec preserved.
    viz_step = steps[2]
    assert viz_step["params"]["chart_type"] == "bar"

    # 6. Drift check: just saved → no drift.
    drift = asset_svc.metadata_drift_check(asset_id)
    assert drift["is_stale"] is False


# ============================================================
# Test 12 (spec §14.5 #12 + §14.1 frozen baseline rule):
#   schema_driven_baseline_unchanged
#
#   Proxy check — these schema-driven files must (a) compile, (b) import
#   without side-effect failures, (c) still export their critical public
#   entry points. A real baseline run (test_runner.py --domain tflex) is
#   in scripts/run_regression_gate.py --mode full (skipped without Ollama).
# ============================================================
def test_multisheet_12_schema_driven_baseline_unchanged():
    import importlib
    import py_compile

    # Critical schema-driven files — these are the ones the v0.18 upload
    # work must NOT have modified in any breaking way. py_compile catches
    # syntax breaks; the existence check catches accidentally renamed/
    # deleted entry points.
    project_root = Path(__file__).resolve().parent.parent.parent
    schema_driven_files = [
        "app.py",
        "llm_service.py",
        "config.py",
        "test_runner.py",
        "pages/main_chat.py",
    ]
    for rel in schema_driven_files:
        p = project_root / rel
        assert p.exists(), f"schema-driven file missing: {rel}"
        py_compile.compile(str(p), doraise=True)

    # Critical schema-driven public entry points still exported. If
    # someone refactored llm_service and removed/renamed any of these,
    # the schema-driven baseline would break — this catches it before a
    # CI test_runner run.
    llm_service = importlib.import_module("llm_service")
    for name in (
        "LLMService",
        "classify_intent",
        "build_domain_knowledge",
        "build_metadata_vocab",
    ):
        assert hasattr(llm_service, name), (
            f"llm_service.{name} disappeared — schema-driven path break"
        )

    # Schema-driven path must not have grown an upload_* import (that
    # would couple the two paths and risk the frozen baseline).
    llm_src = (project_root / "llm_service.py").read_text(encoding="utf-8")
    forbidden = ("from upload_", "import upload_")
    for needle in forbidden:
        assert needle not in llm_src, (
            f"llm_service.py contains `{needle}` — upload-driven code has "
            f"leaked into schema-driven path, violating spec §14.1"
        )
